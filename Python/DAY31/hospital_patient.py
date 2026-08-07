import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter




# Step 1: Read CSV
df_raw = pd.read_csv("E:/OM AI/AI-Engineer-Bootcamp/Python/DAY31/hospital_patients_raw.csv")

# Step 2 & 3: Dataset Info & Missing values analysis
print("=== 1. RAW DATASET ===")
print(df_raw)
print("\n=== 2. DATASET INFO ===")
df_raw.info()
print("\n=== 3. MISSING VALUES PER COLUMN ===")
missing_series = df_raw.isnull().sum()
print(missing_series)

# Step 4: Clean data
df_clean = df_raw.copy()
# Impute missing Bill with median of department
df_clean['Bill'] = df_clean.groupby('Department')['Bill'].transform(lambda x: x.fillna(x.median()))
# Impute missing Age with median overall (or department)
df_clean['Age'] = df_clean['Age'].fillna(df_clean['Age'].median()).astype(int)
# Impute missing StayDays with median stay days
df_clean['StayDays'] = df_clean['StayDays'].fillna(df_clean['StayDays'].median()).astype(int)

print("\n=== 4. CLEANED DATASET ===")
print(df_clean)

# Step 5: Department-wise bill & summary metrics
dept_summary = df_clean.groupby('Department', as_index=False).agg(
    Total_Patients=('PatientID', 'count'),
    Total_Bill=('Bill', 'sum'),
    Average_Bill=('Bill', 'mean'),
    Min_Bill=('Bill', 'min'),
    Max_Bill=('Bill', 'max'),
    Avg_Stay_Days=('StayDays', 'mean'),
    Avg_Age=('Age', 'mean')
)

dept_summary['Total_Bill'] = dept_summary['Total_Bill'].round(2)
dept_summary['Average_Bill'] = dept_summary['Average_Bill'].round(2)
dept_summary['Avg_Stay_Days'] = dept_summary['Avg_Stay_Days'].round(1)
dept_summary['Avg_Age'] = dept_summary['Avg_Age'].round(1)

print("\n=== 5. DEPARTMENT-WISE BILL SUMMARY ===")
print(dept_summary.to_string(index=False))

# Step 6: Save clean CSV
df_clean.to_csv("hospital_patients_cleaned.csv", index=False)
print("\n✔ Saved 'hospital_patients_cleaned.csv'")

# Step 7: Export Excel Report with Professional Styling
wb = openpyxl.Workbook()
# Default sheet
ws_summary = wb.active
ws_summary.title = "Department Summary"
ws_patients = wb.create_sheet(title="Cleaned Patient Data")
ws_raw = wb.create_sheet(title="Raw Data (Audit)")

# Styling definitions
font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
font_subtitle = Font(name="Calibri", size=11, italic=True, color="595959")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_bold = Font(name="Calibri", size=11, bold=True)
font_regular = Font(name="Calibri", size=11)

fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
fill_light_blue = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
fill_zebra = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
top_thin_bottom_double = Border(
    top=Side(style='thin', color='000000'),
    bottom=Side(style='double', color='000000')
)

# ----------------- SHEET 1: Summary Dashboard -----------------
ws_summary.views.sheetView[0].showGridLines = True
ws_summary['A1'] = "HOSPITAL PATIENT & BILLING ANALYTICS"
ws_summary['A1'].font = font_title
ws_summary['A2'] = "Departmental Breakdown and Performance Metrics"
ws_summary['A2'].font = font_subtitle

# Write KPI Cards
kpi_labels = ["Total Patients", "Total Revenue", "Avg Bill / Patient", "Avg Length of Stay"]
kpi_values = [
    len(df_clean),
    df_clean['Bill'].sum(),
    df_clean['Bill'].mean(),
    f"{df_clean['StayDays'].mean():.1f} Days"
]

for col_idx, (label, val) in enumerate(zip(kpi_labels, kpi_values), start=1):
    cell_lbl = ws_summary.cell(row=4, column=col_idx, value=label)
    cell_val = ws_summary.cell(row=5, column=col_idx, value=val)
    
    cell_lbl.font = Font(name="Calibri", size=9, bold=True, color="595959")
    cell_lbl.alignment = Alignment(horizontal="center", vertical="center")
    cell_lbl.fill = fill_light_blue
    
    cell_val.font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    cell_val.alignment = Alignment(horizontal="center", vertical="center")
    if isinstance(val, (int, float)):
        cell_val.number_format = '$#,##0.00' if isinstance(val, float) else '#,##0'

# Write Department Table Header
headers_summary = ["Department", "Patient Count", "Total Billing ($)", "Average Bill ($)", "Min Bill ($)", "Max Bill ($)", "Avg Stay (Days)", "Avg Age (Yrs)"]
for col_idx, h in enumerate(headers_summary, start=1):
    cell = ws_summary.cell(row=8, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_navy
    cell.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")

# Populate Department Table Data
for row_idx, row_data in enumerate(dept_summary.itertuples(index=False), start=9):
    ws_summary.cell(row=row_idx, column=1, value=row_data.Department).alignment = Alignment(horizontal="left")
    ws_summary.cell(row=row_idx, column=2, value=row_data.Total_Patients).number_format = '#,##0'
    ws_summary.cell(row=row_idx, column=3, value=row_data.Total_Bill).number_format = '$#,##0.00'
    ws_summary.cell(row=row_idx, column=4, value=row_data.Average_Bill).number_format = '$#,##0.00'
    ws_summary.cell(row=row_idx, column=5, value=row_data.Min_Bill).number_format = '$#,##0.00'
    ws_summary.cell(row=row_idx, column=6, value=row_data.Max_Bill).number_format = '$#,##0.00'
    ws_summary.cell(row=row_idx, column=7, value=row_data.Avg_Stay_Days).number_format = '0.0'
    ws_summary.cell(row=row_idx, column=8, value=row_data.Avg_Age).number_format = '0.0'
    
    for col_idx in range(1, 9):
        c = ws_summary.cell(row=row_idx, column=col_idx)
        c.font = font_regular
        c.border = thin_border
        if row_idx % 2 == 0:
            c.fill = fill_zebra

# Total Row
tot_row = 9 + len(dept_summary)
ws_summary.cell(row=tot_row, column=1, value="Total / Overall").font = font_bold
ws_summary.cell(row=tot_row, column=2, value=f"=SUM(B9:B{tot_row-1})").number_format = '#,##0'
ws_summary.cell(row=tot_row, column=3, value=f"=SUM(C9:C{tot_row-1})").number_format = '$#,##0.00'
ws_summary.cell(row=tot_row, column=4, value=f"=AVERAGE('Cleaned Patient Data'!E2:E{len(df_clean)+1})").number_format = '$#,##0.00'
ws_summary.cell(row=tot_row, column=5, value=f"=MIN('Cleaned Patient Data'!E2:E{len(df_clean)+1})").number_format = '$#,##0.00'
ws_summary.cell(row=tot_row, column=6, value=f"=MAX('Cleaned Patient Data'!E2:E{len(df_clean)+1})").number_format = '$#,##0.00'
ws_summary.cell(row=tot_row, column=7, value=f"=AVERAGE('Cleaned Patient Data'!G2:G{len(df_clean)+1})").number_format = '0.0'
ws_summary.cell(row=tot_row, column=8, value=f"=AVERAGE('Cleaned Patient Data'!F2:F{len(df_clean)+1})").number_format = '0.0'

for col_idx in range(1, 9):
    c = ws_summary.cell(row=tot_row, column=col_idx)
    c.font = font_bold
    c.border = top_thin_bottom_double

# Add Column Chart in Summary Sheet
from openpyxl.chart import BarChart, Reference
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Total Billing by Department ($)"
chart.y_axis.title = "Total Bill ($)"
chart.x_axis.title = "Department"

data_ref = Reference(ws_summary, min_col=3, min_row=8, max_row=8+len(dept_summary))
cats_ref = Reference(ws_summary, min_col=1, min_row=9, max_row=8+len(dept_summary))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 16
chart.height = 10
ws_summary.add_chart(chart, "J4")


# ----------------- SHEET 2: Cleaned Data -----------------
ws_patients.views.sheetView[0].showGridLines = True
headers_patient = list(df_clean.columns)
for col_idx, h in enumerate(headers_patient, start=1):
    cell = ws_patients.cell(row=1, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = fill_navy

for row_idx, row_data in enumerate(df_clean.itertuples(index=False), start=2):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws_patients.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_regular
        cell.border = thin_border
        if headers_patient[col_idx-1] == 'Bill':
            cell.number_format = '$#,##0.00'
        elif headers_patient[col_idx-1] in ['Age', 'StayDays']:
            cell.number_format = '#,##0'

# ----------------- SHEET 3: Raw Data -----------------
ws_raw.views.sheetView[0].showGridLines = True
headers_raw = list(df_raw.columns)
for col_idx, h in enumerate(headers_raw, start=1):
    cell = ws_raw.cell(row=1, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = PatternFill(start_color="595959", end_color="595959", fill_type="solid")

for row_idx, row_data in enumerate(df_raw.itertuples(index=False), start=2):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws_raw.cell(row=row_idx, column=col_idx, value=val)
        cell.font = font_regular
        cell.border = thin_border

# Auto-fit column widths across all sheets
for ws in [ws_summary, ws_patients, ws_raw]:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # convert formula or string
                val_str = str(cell.value)
                if not val_str.startswith("="):
                    max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

# Specific manual adjustment for KPI cards
ws_summary.column_dimensions['A'].width = 18
ws_summary.column_dimensions['B'].width = 16
ws_summary.column_dimensions['C'].width = 18
ws_summary.column_dimensions['D'].width = 18

excel_file = "hospital_analytics_report.xlsx"
wb.save(excel_file)
print(f"✔ Exported Excel Report to '{excel_file}'")

# Step 8: Print statistics summary
total_patients = len(df_clean)
total_revenue = df_clean['Bill'].sum()
avg_bill = df_clean['Bill'].mean()
highest_bill_patient = df_clean.loc[df_clean['Bill'].idxmax()]
longest_stay_patient = df_clean.loc[df_clean['StayDays'].idxmax()]
highest_revenue_dept = dept_summary.loc[dept_summary['Total_Bill'].idxmax()]

print("\n" + "="*50)
print("=== 8. EXECUTIVE HOSPITAL ANALYTICS SUMMARY ===")
print("="*50)
print(f"• Total Patients Analyzed   : {total_patients}")
print(f"• Total Hospital Revenue    : ${total_revenue:,.2f}")
print(f"• Overall Average Bill      : ${avg_bill:,.2f}")
print(f"• Highest Revenue Department: {highest_revenue_dept['Department']} (${highest_revenue_dept['Total_Bill']:,.2f})")
print(f"• Highest Billed Patient    : {highest_bill_patient['Name']} ({highest_bill_patient['Department']}) - ${highest_bill_patient['Bill']:,.2f}")
print(f"• Longest Stay Patient      : {longest_stay_patient['Name']} ({longest_stay_patient['Department']}) - {longest_stay_patient['StayDays']} Days")
print("="*50)